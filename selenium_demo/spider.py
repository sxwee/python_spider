from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
import openpyxl


#初始化driver
driver = webdriver.Chrome()
WAIT = WebDriverWait(driver, 10)

def FirstPage(keywords):
    """
    进入B站待爬取的页面并返回带爬取的页面数
    """
    try:
        #进入B站主页
        driver.get("https://www.bilibili.com/")

        #获取主页的搜索输入框和按钮
        input = WAIT.until(EC.presence_of_element_located((By.XPATH,"//form[@id='nav_searchform']/input[@type='text']")))
        button = WAIT.until(EC.element_to_be_clickable((By.XPATH,"//form[@id='nav_searchform']//button[@type='button']")))

        #在B站主页面输入搜索关键词
        input.send_keys(keywords)
        #点击搜索按钮
        button.click()
        
        print("-----------------正在爬取第1页-----------------")
        #跳转至待爬取页面
        hds = driver.window_handles
        driver.switch_to_window(hds[-1])
        
        return getPages()#返回待爬取的总页数

    except TimeoutException:
        FirstPage()


def getPages():
    """
    获取页面数
    """
    pages = WAIT.until(EC.presence_of_element_located((By.XPATH,"//div[@id='all-list']\
    /div[@class='flow-loader']//ul[@class='pages']//button[@class='pagination-btn']")))
    
    return int(pages.text)

def parseHTML(html):
    """
    解析HTML页面
    """
    vlist = []
    soup = BeautifulSoup(html,'lxml')
    videos = soup.findAll('li',attrs = {"class":"video-item matrix"})
    for video in videos:
        title = video.find('a').get('title')
        href = video.find('a').get('href')[2:]
        times = video.find('span',attrs = {"class":"so-icon watch-num"}).text.strip()
        barrage = video.find('span',attrs = {"class":"so-icon hide"}).text.strip()
        print(title,href,times,barrage)
        vlist.append([title,href,times,barrage])
    return vlist
    

def getNextPage(page_num):
    """
    page_nums为获取到的页面数
    获取下一个页面
    """
    try:
        next_button = WAIT.until(EC.element_to_be_clickable((By.CSS_SELECTOR,"#all-list > div.flow-loader > div.page-wrap >\
                 div > ul > li.page-item.next > button"
        )))
        next_button.click()
        WAIT.until(EC.text_to_be_present_in_element((By.CSS_SELECTOR, '#all-list > div.flow-loader > div.page-wrap > div > ul > li.page-item.active > button'),str(page_num)))
        html = driver.page_source
        return parseHTML(html)
    except TimeoutException:
        driver.refresh()
        return getNextPage(page_num)

def main(keywords):
    """
    爬取关键词在B站的所有视频信息
    返回值为一个包含所有信息的列表
    """
    #保存最终结果的列表
    vlists = []

    #获取第一个页面
    page_nums = FirstPage(keywords)
    first_page = driver.page_source
    vlist = parseHTML(first_page)

    vlists += vlist

    i = 2
    try:
        while i < page_nums + 1:
            print("-----------------正在爬取第{}页-----------------".format(i))
            vlist = getNextPage(i)
            vlists += vlist
            i += 1
    finally:
        driver.close()
    
    return vlists

def Saver(vlist):
    """
    vlist：视频信息列表
    将数据保存在excel中
    """        
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = '爬虫结果'
    sheet.cell(1,1,'视频名')
    sheet.cell(1,2,'视频链接')
    sheet.cell(1,3,'播放量')
    sheet.cell(1,4,'弹幕数')

    for idx,m in enumerate(vlist):
        sheet.cell(idx + 2,1,m[0])
        sheet.cell(idx + 2,2,m[1])
        sheet.cell(idx + 2,3,m[2])
        sheet.cell(idx + 2,4,m[3])

    workbook.save(u'JianGuo_info.xlsx')  # 保存工作簿

if __name__ == "__main__":
    keywords = "川建国"
    vlists = main(keywords)
    Saver(vlists)

